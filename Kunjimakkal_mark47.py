# Importing Necessary Modules
import kunjiconfig
from googleapiclient.discovery import build
from httplib2 import Http
from oauth2client import file, client, tools
from datetime import date
from datetime import datetime
from openpyxl import Workbook

# Scopes of the API as defined by google
SCOPES = 'https://www.googleapis.com/auth/gmail.readonly'

def main():
    # Initializing Excel Sheet
    row_no = 2
    book = Workbook()
    sheet = book.active
    try:
        sheet.cell(row=1, column=1).value = "DATE"
        sheet.cell(row=1, column=2).value = "MORNING"
        sheet.cell(row=1, column=3).value = "AFTERNOON"
        sheet.cell(row=1, column=4).value = "NIGHT"
        book.save("funds2.xlsx")

    
    except :
        pass

    # Storing the tokens and credentials.
    # token.json contains authorised token of user
    store = file.Storage("token.json")
    creds = None
    if not creds or creds.invalid:
        # credentials.json is obtained from google cloud 
        flow = client.flow_from_clientsecrets('credentials.json', SCOPES)
        creds = tools.run_flow(flow, store)
   
   
    service = build('gmail', 'v1', http=creds.authorize(Http()))
    

    breakfast = []
    lunch = []
    dinner = []
    lunch_expense = 0
    dinner_expense = 0 
    breakfast_expense = 0
    
    # sender is the user id of the service providing transaction details (Here it is 'IndusInd_Bank@indusind.com' )
    query = f"from : {kunjiconfig.sender}"

    sheet.cell(row_no, column=1).value = str(date.today())
    book.save("funds2.xlsx")

    while True:

        # Retrieving messages from a given sender
        results = service.users().messages().list(userId='me',labelIds = ['INBOX'] , q = query).execute()
        messages = results.get('messages', [])

        # Getting the present time
        now = datetime.now()
        
        # To check if it is the next day and initializing the expenses for the next day
        if now.hour == 00 and now.minute == 00:
            row_no += 1
            lunch_expense = 0
            dinner_expense = 0 
            breakfast_expense = 0
            sheet.cell(row_no, column=1).value = str(date.today())
            book.save("funds2.xlsx")

        

        if not messages:
            print ("No messages found.")
        else:
            
            for message in messages:
                
                msg = service.users().messages().get(userId='me', id=message['id']).execute()

                splittd_msg = msg['snippet'].split(" ")
                # Fetching index of the expense from the list containing the splitted string
                indeX = splittd_msg.index("INR") +1
               
                # Checking whether if keywords ('brea','lun','dinn') are present in the message and updating corresponding
                # parameters in the excel sheet
                try:
                    if 'brea' in msg['snippet'] and (msg['snippet'] not in breakfast) :
                        breakfast_expense += float(splittd_msg[indeX])
                        sheet.cell(row_no, column=2).value = breakfast_expense
                        book.save("funds2.xlsx")
                        breakfast.append(msg['snippet'])

                    if 'lun' in msg['snippet'] and (msg['snippet'] not in lunch) :
                        lunch_expense += float(splittd_msg[indeX])
                        sheet.cell(row_no, column=3).value = lunch_expense
                        book.save("funds2.xlsx")
                        lunch.append(msg['snippet'])
                    
                    if 'dinn' in msg['snippet'] and (msg['snippet'] not in dinner) :
                        dinner_expense += float(splittd_msg[indeX])
                        sheet.cell(row_no, column=4).value = dinner_expense
                        book.save("funds2.xlsx")
                        dinner.append(msg['snippet'])
                
                except:
                    pass
                
# Executing the main function
if __name__ == '__main__':
    main()